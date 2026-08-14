#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""呆滞批次周报（每周五 16:00 北京时间推送钉钉群机器人）。

流程：
  1. 读取 data/batches/initial.json（初始批次快照）+ data/batches/name-map.json（全名→精简名）
     + data/records/*.json（出入库记录）
  2. 推导批次台账（与前端 src/js/data/batch.js 同构）：
       批次当前库存 = initial 快照 + Σ(in 带 batchNo) - Σ(out batchAlloc 扣减)
       旧出库记录（无 batchAlloc）不扣批次
  3. openpyxl 生成《产品批号库存库龄汇总_深圳细胞时空仓_YYYYMMDD.xlsx》（两 sheet：库存台账+产品维度汇总）
  4. 上传到 data/reports/ → jsdelivr 公网 URL
  5. markdown 卡片：KPI + 呆滞批次列表 + 批次台账摘要表 + xlsx 下载链接 → 钉钉

读取环境变量：
  WEBHOOK   : 钉钉群机器人 Webhook 地址
  SECRET    : 钉钉安全设置「加签」密钥
  GH_TOKEN  : 仓库令牌（上传 xlsx 到 data/reports/）
"""
import base64
import hashlib
import hmac
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

WEBHOOK = os.environ.get("WEBHOOK", "").strip()
SECRET = os.environ.get("SECRET", "").strip()
GH_TOKEN = os.environ.get("GH_TOKEN", "").strip()

CST = timezone(timedelta(hours=8))

INITIAL_PATH = "data/batches/initial.json"
NAME_MAP_PATH = "data/batches/name-map.json"
RECORDS_DIR = "data/records"
REPORTS_DIR = "data/reports"
GH_REPO = "chenliguan42057/outbound-registry"
GH_BRANCH = "main"

SLUGGISH_DAYS = 180   # 6 个月（呆滞线）
MID_DAYS = 90         # 3 个月（三档分界）


def sign_url(webhook, secret):
    timestamp = str(round(time.time() * 1000))
    string_to_sign = "{}\n{}".format(timestamp, secret)
    hmac_code = hmac.new(
        secret.encode("utf-8"), string_to_sign.encode("utf-8"), digestmod=hashlib.sha256
    ).digest()
    sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
    return webhook + "&timestamp=" + timestamp + "&sign=" + sign


def load_json(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError) as exc:
        print("SKIP {}: {}".format(path, exc))
        return None


def load_records():
    recs = []
    try:
        files = os.listdir(RECORDS_DIR)
    except OSError:
        return recs
    for fn in files:
        if not fn.endswith(".json"):
            continue
        data = load_json(os.path.join(RECORDS_DIR, fn))
        if data is not None:
            recs.append(data)
    return recs


def parse_dt(v):
    """解析 "YYYY-MM-DD HH:MM:SS" / "YYYY-MM-DDTHH:MM" → naive datetime（按北京时间）。"""
    if not v:
        return None
    s = str(v).strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19], fmt)
        except ValueError:
            continue
    return None


def age_days(in_time):
    d = parse_dt(in_time)
    if not d:
        return 0
    now = datetime.now(CST).replace(tzinfo=None)
    return max(0, (now - d).days)


def exp_left_days(exp_date):
    d = parse_dt(exp_date)
    if not d:
        return 0
    now = datetime.now(CST).replace(tzinfo=None)
    return (d - now).days


def sluggish_label(age):
    if age > SLUGGISH_DAYS:
        return "6个月以上"
    if age >= MID_DAYS:
        return "3-6个月"
    return "3个月以内"


def build_ledger(initial, records, name_map):
    """推导批次台账，返回 [dict]（按库龄降序→入库时间升序），与前端 getLedger 同构。"""
    rows = {}
    def norm(n):
        if name_map and name_map.get(n):
            return name_map[n]
        return n
    # 1) 初始批次快照
    for b in (initial or {}).get("batches") or []:
        key = norm(b.get("name", "")) + "|" + str(b.get("batchNo", ""))
        rows[key] = {
            "name": norm(b.get("name", "")), "batchNo": str(b.get("batchNo", "")),
            "qty": int(b.get("qty") or 0), "unit": (initial or {}).get("unit") or "件",
            "inTime": str(b.get("inTime") or ""), "expDate": str(b.get("expDate") or ""),
            "prodDate": str(b.get("prodDate") or ""),
        }
    # 2) 入库记录（带批次字段）
    for rec in records:
        if str(rec.get("type", "")).lower() != "in" or not rec.get("affectsStock"):
            continue
        for it in rec.get("items") or []:
            bno = it.get("batchNo")
            if not bno:
                continue
            key = norm(it.get("name", "")) + "|" + str(bno)
            row = rows.get(key)
            if row:
                row["qty"] += int(it.get("qty") or 0)
            else:
                rows[key] = {
                    "name": norm(it.get("name", "")), "batchNo": str(bno),
                    "qty": int(it.get("qty") or 0), "unit": "件",
                    "inTime": str(rec.get("time") or ""), "expDate": str(it.get("expDate") or ""),
                    "prodDate": str(it.get("prodDate") or ""),
                }
    # 3) 出库记录（带 batchAlloc）按批扣减
    for rec in records:
        if str(rec.get("type", "")).lower() == "in" or not rec.get("affectsStock"):
            continue
        for it in rec.get("items") or []:
            for al in it.get("batchAlloc") or []:
                key = norm(it.get("name", "")) + "|" + str(al.get("batchNo", ""))
                if key in rows:
                    rows[key]["qty"] -= int(al.get("qty") or 0)
    # 4) 过滤 + 派生 + 排序（库龄降序 → 入库时间升序，风险批次在前）
    out = []
    for row in rows.values():
        if row["qty"] <= 0:
            continue
        age = age_days(row["inTime"])
        row["ageDays"] = age
        row["expLeftDays"] = exp_left_days(row["expDate"])
        row["sluggish"] = sluggish_label(age)
        out.append(row)
    order = {"6个月以上": 0, "3-6个月": 1, "3个月以内": 2}
    out.sort(key=lambda r: (order.get(r["sluggish"], 9), -r["ageDays"], r["inTime"]))
    return out


def to_excel_rows(ledger, warehouse, unit):
    """两 sheet 行数据（台账按产品名分组排序，导出可合并产品名列）。"""
    ledger_rows = [[warehouse, "产品名称", "生产批号", "库存数量", "呆滞预警", "单位", "入库时间", "库龄(天)", "到期时间", "剩余天数"]]
    for r in sorted(ledger, key=lambda x: (x["name"], x["inTime"])):
        ledger_rows.append([warehouse, r["name"], r["batchNo"], r["qty"], r["sluggish"], unit,
                            r["inTime"], r["ageDays"], r["expDate"], r["expLeftDays"]])
    # 产品维度汇总
    agg = {}
    for r in ledger:
        s = agg.setdefault(r["name"], {"name": r["name"], "batchCount": 0, "totalQty": 0, "earliestIn": "", "nearestExp": ""})
        s["batchCount"] += 1
        s["totalQty"] += r["qty"]
        if not s["earliestIn"] or r["inTime"] < s["earliestIn"]:
            s["earliestIn"] = r["inTime"]
        if not s["nearestExp"] or (r["expDate"] and r["expDate"] < s["nearestExp"]):
            s["nearestExp"] = r["expDate"]
    summary_rows = [[warehouse, "产品名称", "批号数", "库存总数量", "最早入库时间", "最近到期时间"]]
    for s in sorted(agg.values(), key=lambda x: x["name"]):
        summary_rows.append([warehouse, s["name"], s["batchCount"], s["totalQty"], s["earliestIn"], s["nearestExp"]])
    return ledger_rows, summary_rows


def gen_xlsx(ledger_rows, summary_rows, out_path):
    """openpyxl 生成两 sheet xlsx（产品名列合并，参考吉客云导出格式）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "库存台账"
    for row in ledger_rows:
        ws1.append(row)
    # 产品名列（B，index 2）合并：同产品连续行
    start = 2
    for i in range(3, len(ledger_rows) + 2):
        cur = ledger_rows[i - 1][1] if i - 1 < len(ledger_rows) else None
        prev = ledger_rows[i - 2][1] if i - 2 < len(ledger_rows) else None
        if i == len(ledger_rows) + 1 or cur != prev:
            if i - 1 > start:
                ws1.merge_cells(start_row=start, start_column=2, end_row=i - 1, end_column=2)
            start = i
    ws1.freeze_panes = "A2"
    widths1 = [16, 34, 20, 9, 10, 6, 20, 9, 20, 9]
    for idx, w in enumerate(widths1, 1):
        ws1.column_dimensions[chr(64 + idx)].width = w
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    ws2 = wb.create_sheet("产品维度汇总")
    for row in summary_rows:
        ws2.append(row)
    widths2 = [16, 34, 8, 12, 20, 20]
    for idx, w in enumerate(widths2, 1):
        ws2.column_dimensions[chr(64 + idx)].width = w
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    wb.save(out_path)
    return True


def upload_xlsx(out_path):
    """上传 xlsx 到 data/reports/，返回 jsdelivr URL；失败返回 None。"""
    if not GH_TOKEN or not os.path.exists(out_path):
        return None
    fname = os.path.basename(out_path)
    path = "{}/{}".format(REPORTS_DIR, fname)
    url = "https://api.github.com/repos/{}/contents/{}".format(GH_REPO, path)
    headers = {
        "Accept": "application/vnd.github+json",
        "Authorization": "Bearer {}".format(GH_TOKEN),
        "X-GitHub-Api-Version": "2022-11-28",
    }
    sha = None
    try:
        req = urllib.request.Request(url, headers=headers, method="GET")
        with urllib.request.urlopen(req, timeout=15) as resp:
            sha = json.loads(resp.read().decode("utf-8")).get("sha")
    except urllib.error.HTTPError as exc:
        if exc.code != 404:
            print("WARN 查询 xlsx {} 失败：{}".format(path, exc), file=sys.stderr)
    except Exception as exc:
        print("WARN 查询 xlsx 失败：{}".format(exc), file=sys.stderr)
    with open(out_path, "rb") as f:
        content = base64.b64encode(f.read()).decode("utf-8")
    body = {"message": "batch report {}".format(fname), "content": content, "branch": GH_BRANCH}
    if sha:
        body["sha"] = sha
    try:
        req = urllib.request.Request(
            url, data=json.dumps(body).encode("utf-8"),
            headers=dict(headers, **{"Content-Type": "application/json"}), method="PUT")
        with urllib.request.urlopen(req, timeout=30) as resp:
            resp.read()
        return "https://cdn.jsdelivr.net/gh/{}@{}/{}/{}".format(GH_REPO, GH_BRANCH, REPORTS_DIR, fname)
    except Exception as exc:
        print("WARN 上传 xlsx {} 失败：{}".format(path, exc), file=sys.stderr)
        return None


def build_markdown(ledger, xlsx_url, warehouse, unit):
    now = datetime.now(CST)
    total_qty = sum(r["qty"] for r in ledger)
    sluggish = [r for r in ledger if r["sluggish"] == "6个月以上"]
    mid = [r for r in ledger if r["sluggish"] == "3-6个月"]
    lines = [
        "### 📦 出入库登记 · 呆滞批次周报",
        "- **生成时间**：{}（北京时间）".format(now.strftime("%Y-%m-%d %H:%M")),
        "- **仓库**：{} ｜ **批次总数**：{} ｜ **在库总数量**：{} {}".format(warehouse, len(ledger), total_qty, unit),
        "- **呆滞批次（>180天）**：**{}** 条 ｜ **3-6个月**：{} 条".format(len(sluggish), len(mid)),
    ]
    # 呆滞批次明细
    lines.append("")
    if sluggish:
        lines.append("**🔴 呆滞批次（库龄>180天，共 {} 条）：**".format(len(sluggish)))
        lines.append("| 产品 | 批号 | 数量 | 入库时间 | 库龄(天) | 剩余天数 |")
        lines.append("| --- | --- | ---: | --- | ---: | ---: |")
        for r in sluggish[:15]:
            lines.append("| {} | {} | {} | {} | **{}** | {} |".format(
                r["name"], r["batchNo"], r["qty"], r["inTime"], r["ageDays"], r["expLeftDays"]))
        if len(sluggish) > 15:
            lines.append("_另有 {} 条呆滞批次见完整 xlsx_".format(len(sluggish) - 15))
    else:
        lines.append("✅ 暂无呆滞批次（库龄>180天）")
    # 台账摘要（风险排序：呆滞 → 3-6个月 → 3个月以内，前 15 条）
    lines.append("")
    lines.append("**📋 批次台账（风险优先，前 15 条）：**")
    lines.append("| 产品 | 批号 | 数量 | 呆滞预警 | 入库时间 | 库龄(天) | 到期时间 | 剩余天数 |")
    lines.append("| --- | --- | ---: | --- | --- | ---: | --- | ---: |")
    for r in ledger[:15]:
        lines.append("| {} | {} | {} | {} | {} | {} | {} | {} |".format(
            r["name"], r["batchNo"], r["qty"], r["sluggish"], r["inTime"], r["ageDays"], r["expDate"], r["expLeftDays"]))
    if len(ledger) > 15:
        lines.append("_另有 {} 条批次见完整 xlsx_".format(len(ledger) - 15))
    # xlsx 下载
    if xlsx_url:
        fname = xlsx_url.split("/")[-1]
        lines.append("")
        lines.append("📥 **完整表格**：[{}]({})".format(fname, xlsx_url))
    lines.append("")
    lines.append("— 每周五 16:00 自动推送 · 数据实时来自云端登记")
    return "\n".join(lines)


def send(text, title="呆滞批次周报"):
    if not WEBHOOK or not SECRET:
        print("WEBHOOK/SECRET 未配置", file=sys.stderr)
        return False
    from ding_card import send_action_card, REG_URL
    return send_action_card(
        text, title, WEBHOOK, SECRET,
        btns=[{"title": "🌿 打开出库登记", "url": REG_URL}],
        btn_orientation="0",
    )


def main():
    initial = load_json(INITIAL_PATH)
    name_map_data = load_json(NAME_MAP_PATH)
    name_map = (name_map_data or {}).get("nameMap") or {}
    records = load_records()

    warehouse = (initial or {}).get("warehouse") or "深圳细胞-时空仓"
    unit = (initial or {}).get("unit") or "件"
    ledger = build_ledger(initial, records, name_map)

    ymd = datetime.now(CST).strftime("%Y%m%d")
    fname = "产品批号库存库龄汇总_深圳细胞时空仓_{}.xlsx".format(ymd)
    out_path = "/tmp/" + fname
    ledger_rows, summary_rows = to_excel_rows(ledger, warehouse, unit)
    try:
        gen_xlsx(ledger_rows, summary_rows, out_path)
        print("xlsx 生成成功: {} 行台账 / {} 行汇总".format(len(ledger_rows) - 1, len(summary_rows) - 1))
    except Exception as exc:
        print("xlsx 生成失败: {}".format(exc), file=sys.stderr)
        out_path = None

    xlsx_url = upload_xlsx(out_path) if out_path else None

    md = build_markdown(ledger, xlsx_url, warehouse, unit)
    print("--- 报告（前 300 字）---")
    print(md[:300])
    if send(md):
        print("呆滞批次周报发送成功")
        return 0
    print("呆滞批次周报发送失败", file=sys.stderr)
    return 1


if __name__ == "__main__":
    sys.exit(main())
